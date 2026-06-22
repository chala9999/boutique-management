import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.http import HttpResponse
from django.utils import timezone

class ExportService:
    
    @staticmethod
    def export_ventes_excel(data):
        """Export des ventes vers Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport des ventes"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Ligne des en-têtes
        headers = ['Date', 'N° Vente', 'Boutique', 'Vendeur', 'Client', 'Montant', 'Mode paiement', 'Statut', 'Statut paiement']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Données
        for row, vente in enumerate(data, 2):
            ws.cell(row=row, column=1, value=vente.date_vente.strftime('%d/%m/%Y %H:%M') if hasattr(vente, 'date_vente') else '')
            ws.cell(row=row, column=2, value=vente.numero_vente if hasattr(vente, 'numero_vente') else '')
            ws.cell(row=row, column=3, value=vente.boutique.nom if hasattr(vente, 'boutique') else '')
            ws.cell(row=row, column=4, value=f"{vente.vendeur.first_name} {vente.vendeur.last_name}" if hasattr(vente, 'vendeur') else '')
            ws.cell(row=row, column=5, value=vente.client.nom_complet if hasattr(vente, 'client') and vente.client else 'Client anonyme')
            ws.cell(row=row, column=6, value=float(vente.montant_final) if hasattr(vente, 'montant_final') else 0)
            ws.cell(row=row, column=7, value=dict(vente.MODE_PAIEMENT_CHOICES).get(vente.mode_paiement, '') if hasattr(vente, 'mode_paiement') else '')
            ws.cell(row=row, column=8, value=dict(vente.STATUT_CHOICES).get(vente.statut, '') if hasattr(vente, 'statut') else '')
            ws.cell(row=row, column=9, value=dict(vente.STATUT_PAIEMENT_CHOICES).get(vente.statut_paiement, '') if hasattr(vente, 'statut_paiement') else '')
        
        # Ajuster les largeurs
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Sauvegarder
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_produits_excel(data):
        """Export des produits vers Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport des stocks"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        
        headers = ['Produit', 'Référence', 'Catégorie', 'Boutique', 'Prix achat', 'Prix vente', 'Stock actuel', 'Stock min', 'Statut', 'Stock faible']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row, produit in enumerate(data, 2):
            ws.cell(row=row, column=1, value=produit.nom)
            ws.cell(row=row, column=2, value=produit.reference)
            ws.cell(row=row, column=3, value=produit.categorie.nom if produit.categorie else '-')
            ws.cell(row=row, column=4, value=produit.boutique.nom)
            ws.cell(row=row, column=5, value=float(produit.prix_achat))
            ws.cell(row=row, column=6, value=float(produit.prix_vente))
            ws.cell(row=row, column=7, value=int(produit.stock_actuel))
            ws.cell(row=row, column=8, value=int(produit.stock_minimum))
            ws.cell(row=row, column=9, value='Actif' if produit.is_active else 'Inactif')
            ws.cell(row=row, column=10, value='Oui' if produit.stock_faible else 'Non')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_finances_excel(data):
        """Export des finances vers Excel"""
        wb = Workbook()
        
        # Feuille CA par mois
        ws_ca = wb.active
        ws_ca.title = "CA par mois"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        
        headers = ['Mois', 'Chiffre d\'affaires', 'Nombre ventes', 'Bénéfice']
        for col, header in enumerate(headers, 1):
            cell = ws_ca.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row, mois in enumerate(data.get('ca_par_mois', []), 2):
            ws_ca.cell(row=row, column=1, value=mois.get('nom_mois', ''))
            ws_ca.cell(row=row, column=2, value=float(mois.get('ca', 0)))
            ws_ca.cell(row=row, column=3, value=int(mois.get('nombre_ventes', 0)))
            ws_ca.cell(row=row, column=4, value=float(mois.get('benefice', 0)))
        
        # Feuille top clients
        ws_clients = wb.create_sheet("Top clients")
        headers_clients = ['Client', 'Téléphone', 'Nombre achats', 'Total dépensé']
        for col, header in enumerate(headers_clients, 1):
            cell = ws_clients.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, client in enumerate(data.get('top_clients', []), 2):
            ws_clients.cell(row=row, column=1, value=f"{client.get('client__prenom', '')} {client.get('client__nom', '')}")
            ws_clients.cell(row=row, column=2, value=client.get('client__telephone', '-'))
            ws_clients.cell(row=row, column=3, value=int(client.get('nombre_achats', 0)))
            ws_clients.cell(row=row, column=4, value=float(client.get('total_achats', 0)))
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_vendeurs_excel(data):
        """Export des performances vendeurs vers Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Performance vendeurs"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        
        headers = ['Classement', 'Vendeur', 'Username', 'Nombre ventes', 'Chiffre d\'affaires', 'Vente moyenne', 'Bénéfice']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row, vendeur in enumerate(data.get('performance', []), 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=vendeur.get('vendeur_nom', ''))
            ws.cell(row=row, column=3, value=vendeur.get('username', ''))
            ws.cell(row=row, column=4, value=int(vendeur.get('total_ventes', 0)))
            ws.cell(row=row, column=5, value=float(vendeur.get('chiffre_affaires', 0)))
            ws.cell(row=row, column=6, value=float(vendeur.get('vente_moyenne', 0)))
            ws.cell(row=row, column=7, value=float(vendeur.get('benefice', 0)))
        
        # Mise en forme des montants
        for col in [5, 6, 7]:
            for row in range(2, len(data.get('performance', [])) + 2):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0.00 "FCFA"'
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output